"""
Конвертация all_advs.json -> all_advs.xlsx.

- Только полезные поля, с русскими заголовками.
- ЧИСЛОВЫЕ ID (walls_material_type_id, deal_type_id и т.п.) выкинуты,
  т.к. без справочника они бесполезны.
- Телефоны из phone_list склеиваются в строку.
- external_seller_2 (markdown [текст](url)) распарсен на имя + ссылку.
- URL-колонки — кликабельные гиперссылки.
- Цена форматируется как "12\u00a0617\u00a0800\u00a0₽", даты — как datetime.
- Заголовок закреплён, включён автофильтр, колонки с адекватной шириной.
"""

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = Path(__file__).parent / "all_advs.json"
DEFAULT_OUTPUT = Path(__file__).parent / "all_advs.xlsx"


# Описание колонок:
# (ключ_источника, заголовок, тип, ширина)
# Ключи с префиксом "__" — вычисляемые (см. compute_value).
# Типы: str | int | float | price | bool | date | datetime | url
COLUMNS: list[tuple[str, str, str, int]] = [
    ("external_id",                     "ID объявления",        "str",      14),
    ("external_url",                    "Ссылка",               "url",      38),
    ("address",                         "Адрес",                "str",      50),
    ("geo_cache_region_name",           "Регион",               "str",      14),
    ("geo_cache_district_name",         "Район",                "str",      22),
    ("geo_cache_micro_district_name",   "Микрорайон",           "str",      22),
    ("geo_cache_settlement_name",       "Поселение",            "str",      22),
    ("geo_cache_housing_complex_name",  "ЖК",                   "str",      32),
    ("geo_cache_street_name",           "Улица",                "str",      26),
    ("geo_cache_building_name",         "Дом",                  "str",      10),

    ("geo_cache_subway_station_name_1", "Метро 1",              "str",      26),
    ("walking_access_1",                "Пешком, мин",          "int",      11),
    ("transport_access_1",              "Транспорт, мин",       "int",      13),
    ("geo_cache_subway_station_name_2", "Метро 2",              "str",      26),
    ("walking_access_2",                "Пешком 2, мин",        "int",      11),
    ("transport_access_2",              "Транспорт 2, мин",     "int",      13),
    ("geo_cache_subway_station_name_3", "Метро 3",              "str",      26),
    ("walking_access_3",                "Пешком 3, мин",        "int",      11),
    ("transport_access_3",              "Транспорт 3, мин",     "int",      13),

    ("built_year",                      "Год постройки",        "int",      13),
    ("storeys_count",                   "Этажность дома",       "int",      13),
    ("storey",                          "Этаж",                 "int",      7),
    ("total_room_count",                "Комнат",               "int",      8),
    ("is_studio",                       "Студия",               "bool",     8),
    ("is_free_planning",                "Своб. планировка",     "bool",     14),
    ("is_new_building",                 "Новостройка",          "bool",     12),

    ("total_square",                    "Общая, м²",            "float",    10),
    ("life_square",                     "Жилая, м²",            "float",    10),
    ("kitchen_square",                  "Кухня, м²",            "float",    10),
    ("ceiling_height",                  "Потолки, м",           "float",    11),

    ("price_rub",                       "Цена, ₽",              "price",    17),
    ("meter_price_rub",                 "Цена за м², ₽",        "price",    15),
    ("price_change_date",               "Изменение цены",       "date",     15),
    ("sale_type_name",                  "Тип продажи",          "str",      18),

    ("__seller_name",                   "Продавец",             "str",      35),
    ("__seller_url",                    "Ссылка на продавца",   "url",      30),
    ("__broker_name",                   "Брокер",               "str",      30),

    ("__phones_main",                   "Телефоны",             "str",      30),
    ("phone_list_xz",                   "Доп. телефоны",        "str",      40),

    ("photo_count",                     "Фото",                 "int",      7),
    ("video_count",                     "Видео",                "int",      7),
    ("__video_url",                     "Ссылка на видео",      "url",      30),

    ("pub_datetime",                    "Опубликовано",         "datetime", 18),
    ("creation_datetime",               "Создано",              "datetime", 18),
    ("offer_pub_duration",              "Дней на сайте",        "int",      13),
    ("media_name",                      "Источник",             "str",      12),
]


_MARKDOWN_LINK_RE = re.compile(r"^\[(?P<text>[^\]]+)\]\((?P<url>[^)]+)\)\s*$")


def parse_markdown_link(raw: str | None) -> tuple[str | None, str | None]:
    """Разбирает строку вида '[текст](url)' на (текст, url)."""
    if not raw or not isinstance(raw, str):
        return None, None
    match = _MARKDOWN_LINK_RE.match(raw.strip())
    if match:
        return match.group("text").strip(), match.group("url").strip()
    return raw, None


def parse_iso_datetime(s: str | None):
    if not s or not isinstance(s, str):
        return None
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.replace(tzinfo=None)
    except ValueError:
        return None


def parse_iso_date(s: str | None):
    if not s or not isinstance(s, str):
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def format_phones(phone_list) -> str | None:
    if not phone_list or not isinstance(phone_list, list):
        return None
    parts: list[str] = []
    for entry in phone_list:
        if isinstance(entry, dict) and "number" in entry:
            num = str(entry["number"])
            if entry.get("is_black"):
                note = entry.get("black_note")
                num = f"{num} [ЧС{': ' + note if note else ''}]"
            parts.append(num)
        elif isinstance(entry, str):
            parts.append(entry)
    return ", ".join(parts) if parts else None


def format_video(value) -> str | None:
    if not value:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        urls: list[str] = []
        for item in value:
            if isinstance(item, str):
                urls.append(item)
            elif isinstance(item, dict):
                url = item.get("url") or item.get("link") or item.get("href")
                if url:
                    urls.append(str(url))
        return urls[0] if urls else None
    return None


def compute_value(item: dict, key: str):
    """Достаёт значение по ключу с учётом вычисляемых полей."""
    if key == "__seller_name":
        return parse_markdown_link(item.get("external_seller_2"))[0]
    if key == "__seller_url":
        return parse_markdown_link(item.get("external_seller_2"))[1]
    if key == "__broker_name":
        broker = item.get("broker") or {}
        return broker.get("short_name") if isinstance(broker, dict) else None
    if key == "__phones_main":
        return format_phones(item.get("phone_list"))
    if key == "__video_url":
        return format_video(item.get("video_list"))
    return item.get(key)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Конвертация JSON объявлений в Excel.",
    )
    parser.add_argument(
        "--input", type=Path, default=DEFAULT_INPUT,
        help=f"Входной JSON (по умолчанию: {DEFAULT_INPUT.name})",
    )
    parser.add_argument(
        "--output", type=Path, default=DEFAULT_OUTPUT,
        help=f"Выходной XLSX (по умолчанию: {DEFAULT_OUTPUT.name})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_file = args.input
    output_file = args.output

    print(f"Читаю {input_file}...")
    with input_file.open("r", encoding="utf-8") as fp:
        data = json.load(fp)
    total = len(data)
    print(f"Загружено объявлений: {total}")

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Объявления")
    ws.freeze_panes = "A2"

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B0B0B0")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    url_font = Font(color="0563C1", underline="single")
    center_align = Alignment(horizontal="center", vertical="center")
    default_align = Alignment(vertical="center", wrap_text=False)

    # Ширины колонок
    for idx, (_, _, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Строка заголовков
    header_row = []
    for _, title, _, _ in COLUMNS:
        cell = WriteOnlyCell(ws, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
        header_row.append(cell)
    ws.append(header_row)

    price_fmt = '#,##0" ₽"'

    # Данные
    for row_idx, item in enumerate(data, start=1):
        row_cells = []
        for key, _, dtype, _ in COLUMNS:
            raw = compute_value(item, key)
            cell = WriteOnlyCell(ws, value=None)
            cell.alignment = default_align

            if raw is None or raw == "":
                cell.value = None
            elif dtype == "int":
                try:
                    cell.value = int(raw)
                    cell.number_format = "0"
                    cell.alignment = center_align
                except (TypeError, ValueError):
                    cell.value = str(raw)
            elif dtype == "float":
                try:
                    cell.value = float(raw)
                    cell.number_format = "0.0"
                    cell.alignment = center_align
                except (TypeError, ValueError):
                    cell.value = str(raw)
            elif dtype == "price":
                try:
                    cell.value = float(raw)
                    cell.number_format = price_fmt
                except (TypeError, ValueError):
                    cell.value = str(raw)
            elif dtype == "datetime":
                dt = parse_iso_datetime(raw) if isinstance(raw, str) else None
                if dt is not None:
                    cell.value = dt
                    cell.number_format = "YYYY-MM-DD HH:MM"
                else:
                    cell.value = str(raw) if raw is not None else None
            elif dtype == "date":
                d = parse_iso_date(raw) if isinstance(raw, str) else None
                if d is not None:
                    cell.value = d
                    cell.number_format = "YYYY-MM-DD"
                else:
                    cell.value = str(raw) if raw is not None else None
            elif dtype == "bool":
                truthy = bool(raw) and raw not in (0, "0", "false", "False")
                cell.value = "Да" if truthy else "Нет"
                cell.alignment = center_align
            elif dtype == "url":
                url = str(raw).strip()
                cell.value = url
                cell.hyperlink = url
                cell.font = url_font
            else:
                cell.value = str(raw)

            row_cells.append(cell)
        ws.append(row_cells)

        if row_idx % 5000 == 0:
            print(f"  записано {row_idx} / {total}")

    # Автофильтр на всю таблицу (включая заголовок)
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{total + 1}"

    print(f"Сохраняю {output_file}...")
    wb.save(output_file)
    size_mb = output_file.stat().st_size / (1024 * 1024)
    print(f"Готово! Размер файла: {size_mb:.1f} MB")


if __name__ == "__main__":
    main()
